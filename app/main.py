from fastapi import FastAPI, Request, File, Form, UploadFile
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from typing import List , Dict
import os
import pandas as pd
import io
import numpy as np
from app.utils import NpEncoder
import openpyxl
from openpyxl.styles import PatternFill, Font

app = FastAPI(title="FastAPI Compare Excel")

# Get the current directory
current_directory = os.path.dirname(os.path.abspath(__file__))
# Get the parent directory
parent_directory = os.path.dirname(current_directory)
# path to the templates directory
templates_directory = os.path.join(parent_directory, "templates")

templates = Jinja2Templates(directory=templates_directory)

@app.get("/")
async def root(request: Request):
    return templates.TemplateResponse("index.html", {"request": request, "page_title": "Home"})

@app.get("/upload")
async def upload_page(request: Request):
    return templates.TemplateResponse("upload.html", {"request": request, "page_title": "Upload Files"})

@app.post("/upload-excel")
async def upload_excel(
    excel_file: UploadFile = File(...),
    school_sheet_names: List[str] = Form(...),
    company_sheet_names: List[str] = Form(...)
):
    
    contents = await excel_file.read()
    # Create a dictionary to store the sheet DataFrames
    dataframes: Dict[str, Dict[str, pd.DataFrame]] = {
        "school": {},
        "company": {}
    }



    school_data = []
    for sheet in school_sheet_names:
        try:
            df = pd.read_excel(io.BytesIO(contents), sheet_name=sheet)
            df = df.replace({np.nan: None})

            # loop thourgh each row and column
            for index, row in df.iterrows():

                if row["SL"] is not None and pd.notna(row["SL"]):
                    sl_value = str(row["SL"]).lower()
                    match sl_value:
                        case s if "cao đẳng khác" in s:
                            # Handle case 1
                            certificate_value = "Cao đẳng khác ngành"

                        case s if "cao đẳng cùng" in s:
                            # Handle case 3
                            certificate_value = "Cao đẳng cùng ngành"
                        case s if "đại học" in s:
                            # Handle case 2
                            certificate_value = "Đại học"
                        case _:
                            pass

                if row["Họ tên"] is not None and pd.notna(row["Họ tên"]):
                    
                    row_data = {
                        "fullName": str(row["Họ tên"] + " " + row["Tên"]).strip(),
                        "Gender": str(row["Giới tính"]).strip(),
                        "dob": str(row["Ngày sinh"]).strip(),
                        "birthPlace": str(row["Nơi sinh"]).strip(),
                        # "phoneNumber": row["SĐT"].replace(" ", "") if pd.notna(row["SĐT"]) else None,
                        # "email": row["Email"],
                        "certificate": str(certificate_value).strip(),
                        "subject": str(row["Ngành"]).strip(),
                        "index": index,
                        "sheet_name": sheet,
                    }
                    school_data.append(row_data)


            dataframes["school"][sheet] = df
        except Exception as e:
            print(f"Error reading sheet {sheet} from school: {e}")


    #read company sheets
    company_data = []
    for sheet in company_sheet_names:
        try:
            df = pd.read_excel(io.BytesIO(contents), sheet_name=sheet)
            df = df.replace({np.nan: None})


            for index, row in df.iterrows():




                if row["Họ và tên"] is not None and pd.notna(row["Họ và tên"]):

                    certificate_value = row["Hệ TN"]
                    # if contain Nghề or nghề then remove it
                    if "nghề" in str(certificate_value).lower() :
                        certificate_value = str(certificate_value).replace("nghề", "").replace("Nghề", "").strip()


                    
                    row_data = {
                        "fullName": str(row["Họ và tên"]).strip(),
                        "Gender": str(row["Giới tính"]).strip(),
                        "dob": str(row["Ngày sinh"]).strip(),
                        "birthPlace": str(row["Nơi sinh"]).strip(),
                        # "phoneNumber": phone_value,
                        # "email": row["Email"],
                        "certificate": str(certificate_value).strip(),
                        "subject": str(row["Ngành ĐK"]).strip(),
                        "index": index,
                        "sheet_name": sheet,
                    }
                    company_data.append(row_data)

            dataframes["company"][sheet] = df
        except Exception as e:
            print(f"Error reading sheet {sheet} from company: {e}")


    wb = openpyxl.load_workbook(io.BytesIO(contents))

    for sheet_name in school_sheet_names:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # find the index of the  "Trạng thái đối chiếu" column
            # check if the column already exists
            compare_col_index = None
            for col in ws.iter_cols(min_row=1, max_row=1):
                for cell in col:
                    if cell.value == "Trạng thái đối chiếu":
                        # if the column already exists then skip
                        compare_col_index = cell.col_idx
                        break
                else:
                    # if the column does not exist then add it
                    # find the last column
                    last_col = ws.max_column + 1
                    # if the first cell is empty then set the first cell to "Trạng thái đối chiếu"
                    ws.cell(row=1, column=last_col, value="Trạng thái đối chiếu")
                    ws.cell(row=1 , column=last_col).font = Font(bold=True)
                    ws.cell(row=1 , column=last_col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    compare_col_index = last_col

            # add data for the new column
            for row_idx in range(2, ws.max_row + 1):
                current_row_data = next(
                    (item for item in school_data if item["index"] == row_idx - 1 and item["sheet_name"] == sheet_name),
                    None
                )
                if current_row_data:
                    # find the current row data exists in company_data
                    match_company_records = [
                        item for item in company_data
                        if str(item["fullName"]).lower() == str(current_row_data["fullName"]).lower() and
                        item["dob"] == current_row_data["dob"]
                    ]

                    match_found = len(match_company_records) > 0

                    if not match_found:
                        ws.cell(row=row_idx, column=compare_col_index, value="Không tìm thấy")
                        ws.cell(row=row_idx, column=compare_col_index).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

                    else:


                        # check if more than 1 record if more than 1 record then the cell value will be yellow color and have orginal value as "Tồn tại nhiều bản ghi"
                        # if only 1 record then compare the other fields
                        if len(match_company_records) == 1:
                            match_company_record = match_company_records[0]
                            # compare the other fields
                            if str(current_row_data["Gender"]).lower() == str(match_company_record["Gender"]).lower() and \
                               str(current_row_data["birthPlace"]).lower() == str(match_company_record["birthPlace"]).lower() and \
                               str(current_row_data["certificate"]).lower() == str(match_company_record["certificate"]).lower() and \
                               str(current_row_data["subject"]).lower() == str(match_company_record["subject"]).lower():
                                ws.cell(row=row_idx, column=last_col, value="Tồn tại và khớp")
                                ws.cell(row=row_idx, column=last_col).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green for match
                            else:
                                # loop through each field to check why not match and current value and company value
                                not_match_fields = []
                                if str(current_row_data["Gender"]).lower() != str(match_company_record["Gender"]).lower():
                                    not_match_fields.append("Giới tính: " + str(current_row_data["Gender"]) + " != " + str(match_company_record["Gender"]))
                                if str(current_row_data["birthPlace"]).lower() != str(match_company_record["birthPlace"]).lower():
                                    not_match_fields.append("Nơi sinh: " + str(current_row_data["birthPlace"]) + " != " + str(match_company_record["birthPlace"]))
                                if str(current_row_data["certificate"]).lower() != str(match_company_record["certificate"]).lower():
                                    not_match_fields.append("Hệ TN: " + str(current_row_data["certificate"]) + " != " + str(match_company_record["certificate"]))
                                if str(current_row_data["subject"]).lower() != str(match_company_record["subject"]).lower():
                                    not_match_fields.append("Ngành ĐK: " + str(current_row_data["subject"]) + " != " + str(match_company_record["subject"]))
                                ws.cell(row=row_idx, column=last_col, value="Không khớp các trường: " + ", ".join(not_match_fields))
                                ws.cell(row=row_idx, column=last_col).fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange for not match

                        else:
                            ws.cell(row=row_idx, column=last_col, value="Tồn tại nhiều bản ghi")
                            ws.cell(row=row_idx, column=last_col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Yellow for multiple records

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    # Create a StreamingResponse to send the modified Excel file
    response = StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={excel_file.filename}"}
    )

    return response

    # return{
    #     "excel_file": excel_file.filename,
    #     "school_sheet_names": school_sheet_names,
    #     "company_sheet_names": company_sheet_names ,
    #     "content_type": excel_file.content_type,
    #     "school_data": school_data,
    #     "company_data": company_data,
    #     "dataframes": {
    #         "school": {sheet: df.to_dict(orient="records") for sheet, df in dataframes["school"].items()},
    #         "company": {sheet: df.to_dict(orient="records") for sheet, df in dataframes["company"].items()}
    #     }
    # }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app.main:app", host="0.0.0.0", port=8000, reload=True)
